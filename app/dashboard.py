# app/dashboard.py
"""
RECAPO Intelligence Dashboard
Real-time collection recovery analytics and forecasting
"""

import sys
from pathlib import Path
from io import BytesIO
#from tkinter import Image
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px

root = Path(__file__).parent.parent
sys.path.insert(0, str(root))

from app.expected_engine import (
    generate_expected_metrics,
    calculate_daily_expected_collection,
    get_collection_summary,
)
from app.collection_engine import calculate_collection_metrics, get_all_contractors_collection
from app.risk_engine import apply_risk_logic, get_risk_distribution, get_portfolio_health_score
from app.recovery_engine import recovery_metrics, get_critical_cases
from app.agent_analytics import contractor_performance, get_top_performers
from app.due_engine import daily_due_customers, get_urgent_followups, get_payment_schedule_summary
from app.forecasting import forecast_recovery, scenario_analysis
from utils.helpers import build_duplicate_phones_table, build_duplicate_names_table

# ================= PAGE CONFIG =================
from PIL import Image
logo = Image.open("data/Recapo Logo.png")
st.set_page_config(
    page_title="RECAPO Intelligence System",
    page_icon=logo,
    layout="wide",
    initial_sidebar_state="expanded"
)

# Add logo
logo_path = Path(__file__).parent.parent / "data" / "Recapo Logo.png"
if logo_path.exists():
    # Add some vertical spacing to move logo down
    st.markdown("<br>", unsafe_allow_html=True)
    
    col1, col2 = st.columns([0.5, 4])
    with col1:
        st.image(str(logo_path), width=100)
    with col2:
        st.title("RECAPO - NNNF")
        st.markdown("Real-time analytics")
else:
    st.title("RECAPO Intelligence System")
    st.markdown("Real-time analytics")

# ================= VISUAL THEME =================
CHART_TEMPLATE = "plotly_white"
CHART_COLORS = {
    "primary": "#15803d",
    "secondary": "#22c55e",
    "accent": "#84cc16",
    "danger": "#dc2626",
    "muted": "#4b6355",
    "surface": "rgba(0,0,0,0)",
    "grid": "#e2e8f0",
}
RISK_COLORS = {
    "On Track": "#16a34a",
    "Watchlist": "#facc15",
    "Medium Risk": "#f97316",
    "High Risk": "#ef4444",
    "Critical": "#7f1d1d",
    "Completed": "#0f766e",
    "Unknown": "#94a3b8",
}

st.markdown(
    """
    <style>
    .block-container {
        padding-top: 1.75rem;
    }
    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        padding: 14px 16px;
        box-shadow: 0 10px 24px rgba(15, 23, 42, 0.06);
    }
    div[data-testid="stMetricLabel"] p {
        color: #475569;
        font-size: 0.82rem;
        font-weight: 650;
    }
    div[data-testid="stMetricValue"] {
        color: #0f172a;
    }
    .stDataFrame {
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        overflow: hidden;
    }
    </style>
    """,
    unsafe_allow_html=True
)


def style_chart(fig, height=380, legend_orientation="h"):
    """Apply a consistent dashboard chart treatment."""
    fig.update_layout(
        template=CHART_TEMPLATE,
        height=height,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=20, t=34, b=28),
        font=dict(family="Arial, sans-serif", size=12, color="#334155"),
        hoverlabel=dict(
            bgcolor="#0f172a",
            bordercolor="#0f172a",
            font=dict(color="#ffffff", size=12)
        ),
        legend=dict(
            orientation=legend_orientation,
            yanchor="bottom",
            y=1.02 if legend_orientation == "h" else 1,
            xanchor="right",
            x=1,
            bgcolor="rgba(0,0,0,0)"
        ),
    )
    fig.update_xaxes(
        showgrid=False,
        linecolor="#d7e3dc",
        tickfont=dict(color="#334155"),
        title_font=dict(color="#1f2937")
    )
    fig.update_yaxes(
        gridcolor="#e2e8f0",
        zerolinecolor="#d7e3dc",
        tickfont=dict(color="#334155"),
        title_font=dict(color="#1f2937")
    )
    return fig


def make_excel_safe(df):
    """Prepare dashboard data for Excel export."""
    export_df = df.copy()
    for col in export_df.columns:
        if pd.api.types.is_datetime64_any_dtype(export_df[col]):
            try:
                export_df[col] = export_df[col].dt.tz_localize(None)
            except TypeError:
                pass
    return export_df


def autosize_sheet(ws):
    for column_cells in ws.columns:
        values = [cell.value for cell in column_cells if cell.value is not None]
        if not values:
            continue
        width = min(max(len(str(value)) for value in values) + 2, 42)
        ws.column_dimensions[column_cells[0].column_letter].width = width
    ws.freeze_panes = "A2"


def build_excel_dashboard(
    filtered_data,
    kpi_data,
    scenario_data,
    contractor_data,
    urgent_data,
    critical_data,
):
    """Create a multi-sheet Excel dashboard workbook with native Excel charts."""
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment

    output = BytesIO()

    risk_summary = (
        filtered_data["Risk_Category"]
        .fillna("Unknown")
        .value_counts()
        .rename_axis("Risk Category")
        .reset_index(name="Customers")
    ) if "Risk_Category" in filtered_data.columns else pd.DataFrame()

    due_export = pd.DataFrame()
    if {"Charged until", "Risk_Category"}.issubset(filtered_data.columns):
        due_source = filtered_data.dropna(subset=["Charged until"]).copy()
        due_source["Due Date"] = due_source["Charged until"].dt.date
        due_export = (
            due_source
            .groupby(["Due Date", "Risk_Category"])
            .size()
            .unstack(fill_value=0)
            .reset_index()
        )
        risk_columns = [col for col in due_export.columns if col != "Due Date"]
        due_export["Total Customers"] = due_export[risk_columns].sum(axis=1)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        kpi_data.to_excel(writer, sheet_name="KPI Summary", index=False)
        risk_summary.to_excel(writer, sheet_name="Risk Summary", index=False)
        due_export.to_excel(writer, sheet_name="Customers by Due Date", index=False)
        scenario_data.to_excel(writer, sheet_name="Scenario Analysis", index=False)
        make_excel_safe(contractor_data).to_excel(writer, sheet_name="Contractors", index=False)
        make_excel_safe(urgent_data).to_excel(writer, sheet_name="Urgent Followups", index=False)
        make_excel_safe(critical_data).to_excel(writer, sheet_name="Critical Cases", index=False)
        make_excel_safe(filtered_data).to_excel(writer, sheet_name="Filtered Data", index=False)

        workbook = writer.book

        for sheet in workbook.worksheets:
            header_fill = PatternFill("solid", fgColor="15803D")
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            autosize_sheet(sheet)

        dashboard = workbook.create_sheet("Excel Dashboard", 0)
        dashboard["A1"] = "RECAPO Excel Dashboard"
        dashboard["A1"].font = Font(size=18, bold=True, color="14532D")
        dashboard["A3"] = "Use the sheets below for KPI summary, risk distribution, due-date load, scenarios, contractors, urgent followups, and critical cases."
        dashboard["A3"].alignment = Alignment(wrap_text=True)
        dashboard.column_dimensions["A"].width = 110

        if len(risk_summary) > 0:
            ws = workbook["Risk Summary"]
            chart = PieChart()
            chart.title = "Risk Distribution"
            labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 8
            chart.width = 12
            dashboard.add_chart(chart, "A5")

        if len(due_export) > 0:
            ws = workbook["Customers by Due Date"]
            max_col = max(ws.max_column - 1, 2)
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.grouping = "stacked"
            chart.overlap = 100
            chart.title = "Customers by Due Date"
            chart.y_axis.title = "Customers"
            chart.x_axis.title = "Due Date"
            data = Reference(ws, min_col=2, max_col=max_col, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 18
            dashboard.add_chart(chart, "J5")

        if len(scenario_data) > 0:
            ws = workbook["Scenario Analysis"]
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Scenario Analysis"
            chart.y_axis.title = "MWK"
            chart.x_axis.title = "Scenario"
            data = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 14
            dashboard.add_chart(chart, "J22")

    output.seek(0)
    return output.getvalue()

# ================= LOAD DATA =================
def load_data():
    """Load latest CSV or Excel export from raw data folder"""
    try:
        files = list((root / "data/raw").glob("*.xlsx")) + list((root / "data/raw").glob("*.csv"))
        if not files:
            raise FileNotFoundError("No CSV or XLSX files found in data/raw")
        latest_file = max(files, key=lambda f: f.stat().st_mtime)
        if latest_file.suffix.lower() == ".csv":
            df = pd.read_csv(latest_file)
        else:
            df = pd.read_excel(latest_file)
        return df
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return pd.DataFrame()

df = load_data()

if len(df) == 0:
    st.error("No data available. Please ensure data/raw/*.xlsx exists.")
    st.stop()

# ================= DATA CLEANING & ENRICHMENT =================
# Standardize column names
df.columns = df.columns.str.strip()

# Clean numeric columns
numeric_cols = ["Balance", "Left to pay", "Days system off", "Payoff amount", "Percentage paid"]
for col in numeric_cols:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

# Standardize state values
if "State" in df.columns:
    df["State"] = df["State"].str.lower().str.strip()

# Extract location if available
location_fields = [
    c for c in df.columns
    if any(key in c.lower() for key in ["location", "area", "zone", "region"])
]
if location_fields:
    df["Location"] = (
        df[location_fields]
        .fillna("")
        .astype(str)
        .agg(", ".join,
             axis=1)
        .str.replace(r"^(, )+|(, )+$", "", regex=True)
        .str.replace(r", ,", ",", regex=True)
        .str.strip()
    )

# ================= ENGINE PIPELINE =================
# Generate expected metrics (core KPI)
df = generate_expected_metrics(df)

# Apply risk logic
df = apply_risk_logic(df)

# ================= SIDEBAR - FILTERS & INFO =================
st.sidebar.header("🎯 Filters & Controls")

# Multi-select contractors
contractors_list = sorted(df["Assigned to contractor"].fillna("Unassigned").astype(str).unique().tolist())
selected_contractors = st.sidebar.multiselect(
    "Select Contractors",
    contractors_list,
    default=contractors_list
)

# Multi-select risk categories
risk_categories = sorted(df["Risk_Category"].fillna("Unknown").astype(str).unique().tolist())
selected_risks = st.sidebar.multiselect(
    "Select Risk Categories",
    risk_categories,
    default=risk_categories
)

# State filter
states = sorted(list(set(df["State"].fillna("Unknown").astype(str).unique().tolist() + ["lead"])))
selected_states = st.sidebar.multiselect(
    "Customer Status",
    states,
    default=states
)

# Location filter
if "Location" in df.columns:
    location_values = sorted(df["Location"].fillna("Unknown").astype(str).unique().tolist())
    selected_locations = st.sidebar.multiselect(
        "Location",
        location_values,
        default=location_values
    )
else:
    selected_locations = None

# Year filter
if "Handover at" in df.columns:
    year_values = sorted(df["Handover at"].dt.year.dropna().astype(int).unique().tolist())
    selected_years = st.sidebar.multiselect(
        "Year",
        year_values,
        default=year_values
    )
else:
    selected_years = None

# Due date and days filters
filter_by_due_date = st.sidebar.checkbox("Filter by Due Date", value=False)
due_date = st.sidebar.date_input(
    "Select Due Date",
    value=pd.Timestamp.now("UTC").date()
)

min_days = int(df["Days_Until_Due"].min()) if "Days_Until_Due" in df.columns else 0
max_days = int(df["Days_Until_Due"].max()) if "Days_Until_Due" in df.columns else 0
selected_days = st.sidebar.slider(
    "Days Until Due Range",
    min_value=min_days,
    max_value=max_days,
    value=(min_days, max_days)
)

if "Days system off" in df.columns:
    system_off_values = (
        pd.to_numeric(df["Days system off"], errors="coerce")
        .fillna(0)
        .clip(lower=0)
        .astype(int)
    )
    max_system_off = int(system_off_values.max()) if len(system_off_values) > 0 else 0
    system_off_filter = st.sidebar.radio(
        "Days System Off",
        ["All", "Range", "Specific days"],
        horizontal=True
    )

    if system_off_filter == "Range":
        selected_system_off_range = st.sidebar.slider(
            "Days System Off Range",
            min_value=0,
            max_value=max_system_off,
            value=(0, max_system_off)
        )
        selected_system_off_days = []
    elif system_off_filter == "Specific days":
        available_system_off_days = sorted(system_off_values.unique().tolist())
        selected_system_off_days = st.sidebar.multiselect(
            "Select Days System Off",
            available_system_off_days,
            default=available_system_off_days
        )
        selected_system_off_range = (0, max_system_off)
    else:
        selected_system_off_range = (0, max_system_off)
        selected_system_off_days = []
else:
    system_off_filter = "All"
    selected_system_off_range = (0, 0)
    selected_system_off_days = []

# Apply filters
contractor_mask = df["Assigned to contractor"].fillna("Unassigned").isin(selected_contractors)
risk_mask = df["Risk_Category"].fillna("Unknown").isin(selected_risks)
state_mask = df["State"].fillna("Unknown").isin(selected_states)

location_mask = pd.Series(True, index=df.index)
if selected_locations is not None:
    location_mask = df["Location"].fillna("Unknown").isin(selected_locations)

days_mask = pd.Series(True, index=df.index)
if selected_days != (min_days, max_days):
    days_mask = df["Days_Until_Due"].between(selected_days[0], selected_days[1])

system_off_mask = pd.Series(True, index=df.index)
if "Days system off" in df.columns:
    system_off_series = (
        pd.to_numeric(df["Days system off"], errors="coerce")
        .fillna(0)
        .clip(lower=0)
        .astype(int)
    )
    if system_off_filter == "Range":
        system_off_mask = system_off_series.between(
            selected_system_off_range[0],
            selected_system_off_range[1]
        )
    elif system_off_filter == "Specific days":
        system_off_mask = system_off_series.isin(selected_system_off_days)

year_mask = pd.Series(True, index=df.index)
if selected_years is not None and "Handover at" in df.columns:
    year_mask = df["Handover at"].dt.year.isin(selected_years)

empty_system_off_selection = system_off_filter == "Specific days" and len(selected_system_off_days) == 0

if len(selected_contractors) == 0 or len(selected_risks) == 0 or len(selected_states) == 0 or empty_system_off_selection or (selected_locations is not None and len(selected_locations) == 0) or (selected_years is not None and len(selected_years) == 0):
    filtered_df = df.iloc[0:0]
else:
    filtered_df = df[contractor_mask & risk_mask & state_mask & days_mask & location_mask & system_off_mask & year_mask]

if filter_by_due_date:
    filtered_df = filtered_df[filtered_df["Charged until"].dt.date == due_date]

# Display filter info
st.sidebar.markdown("---")
st.sidebar.info(f"""
**Data Summary**
- Total records: {len(df)}
- Filtered: {len(filtered_df)}
- Active customers: {len(df[df['State'].isin(['good', 'active'])])}
- Last updated: {df['Last token time'].max() if 'Last token time' in df.columns else 'Unknown'}
""")

# Recalculate metrics after filtering
filtered_df = filtered_df.copy()
daily_collection = calculate_daily_expected_collection(filtered_df)
recovery = recovery_metrics(filtered_df)
collection_metrics = calculate_collection_metrics(filtered_df)
collection_summary = get_collection_summary(filtered_df)
forecast = forecast_recovery(filtered_df, 30)
health_score = get_portfolio_health_score(filtered_df)

# Total loan amount
total_loan = filtered_df['Payoff amount'].sum() if 'Payoff amount' in filtered_df.columns else 0

today_utc = pd.Timestamp.now("UTC").normalize()
if {"Charged until", "Expected_Arrears"}.issubset(filtered_df.columns):
    active_due_mask = pd.Series(True, index=filtered_df.index)
    if "State" in filtered_df.columns:
        active_due_mask = filtered_df["State"].isin(["good", "active"])

    due_and_overdue = filtered_df[
        active_due_mask &
        filtered_df["Charged until"].notna() &
        (filtered_df["Charged until"] <= today_utc)
    ]
    due_and_overdue_expected = due_and_overdue["Expected_Arrears"].sum()
    due_and_overdue_count = len(due_and_overdue)
else:
    due_and_overdue_expected = 0
    due_and_overdue_count = 0

# ================= KEY PERFORMANCE INDICATORS =================
st.markdown("## 📈 Key Performance Indicators")

col1, col2, col3, col4, col5, col6, col7 = st.columns(7)

with col1:
    st.metric(
        "👥 Total Customers",
        f"{len(filtered_df):,}",
        delta=f"{len(df):,} total"
    )

with col2:
    st.metric(
        "💰 Expected Monthly",
        f"MK {collection_summary['monthly_expected']:,.0f}",
        delta=f"{collection_metrics['efficiency_percent']:.1f}% efficiency"
    )

with col3:
    st.metric(
        "📊 Outstanding Balance",
        f"MK {collection_summary['total_outstanding']:,.0f}",
        delta=f"{len(df[df['Risk_Category'].isin(['Critical', 'High Risk'])])} at risk"
    )

with col4:
    st.metric(
        "💰 Total Loan Value",
        f"MK {total_loan:,.0f}",
        delta="Payoff amount total"
    )

with col5:
    st.metric(
        "⏰ Due Today",
        f"{daily_collection['expected_customers']}",
        delta=f"MK {daily_collection['expected_collection']:,.0f}"
    )

with col6:
    st.metric(
        "🏥 Portfolio Health",
        f"{health_score:.0f}/100",
        delta=f"{collection_summary['default_rate']:.1f}% default"
    )

with col7:
    st.metric(
        "⚠️ Default Rate",
        f"{collection_summary['default_rate']:.1f}%",
        delta=f"{daily_collection['default_customers']} default customers"
    )

st.markdown(f"**Total Loan Value:** MK {total_loan:,.0f}  |  **Locations Selected:** {len(selected_locations) if selected_locations is not None else 'N/A'}")

kpi_summary = pd.DataFrame([
    {
        "KPI": "Total Customers",
        "Value": f"{len(filtered_df):,}",
        "Context": f"{len(df):,} records loaded"
    },
    {
        "KPI": "Expected Monthly",
        "Value": f"MK {collection_summary['monthly_expected']:,.0f}",
        "Context": f"{collection_metrics['efficiency_percent']:.1f}% collection efficiency"
    },
    {
        "KPI": "Outstanding",
        "Value": f"MK {collection_summary['total_outstanding']:,.0f}",
        "Context": f"{len(filtered_df[filtered_df['Risk_Category'].isin(['Critical', 'High Risk'])])} high-risk customers"
    },
    {
        "KPI": "Total Loan Value",
        "Value": f"MK {total_loan:,.0f}",
        "Context": "Payoff amount total"
    },
    {
        "KPI": "Due Today",
        "Value": f"{daily_collection['expected_customers']:,}",
        "Context": f"MK {daily_collection['expected_collection']:,.0f} expected"
    },
    {
        "KPI": "Due & Overdue Expected",
        "Value": f"MK {due_and_overdue_expected:,.0f}",
        "Context": f"{due_and_overdue_count:,} customers due today or earlier"
    },
    {
        "KPI": "Portfolio Health",
        "Value": f"{health_score:.0f}/100",
        "Context": f"{collection_summary['default_rate']:.1f}% default rate"
    },
    {
        "KPI": "30-Day Forecast",
        "Value": f"MK {forecast['forecast_30_days']:,.0f}",
        "Context": f"{forecast['confidence_percent']:.0f}% confidence"
    },
])

st.subheader("KPI Summary")
st.dataframe(
    kpi_summary,
    use_container_width=True,
    hide_index=True
)

# Duplicate audit tables
duplicate_phones_df = build_duplicate_phones_table(df)
duplicate_names_df = build_duplicate_names_table(df)

st.markdown("---")
st.subheader("🔍 Duplicate Audit")

# Phone duplicates section
st.markdown("#### Duplicate Phone Numbers")
if len(duplicate_phones_df) > 0:
    st.markdown(
        "Customers with the same phone number. These records may represent the same person or account."
    )
    col1, col2 = st.columns([0.2, 0.8])
    with col1:
        csv_data = duplicate_phones_df.to_csv(index=False)
        st.download_button(
            label="📥 Download (CSV)",
            data=csv_data,
            file_name=f"duplicate_phones_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            key="duplicate_phones_download"
        )
    st.dataframe(duplicate_phones_df, use_container_width=True)
else:
    st.info("No duplicate phone numbers detected in the dataset.")

st.markdown("")

# Name duplicates section
st.markdown("#### Duplicate Customer Names")
if len(duplicate_names_df) > 0:
    st.markdown(
        "Customers with exact or similar names. Review these pairs before processing recoveries."
    )
    col1, col2 = st.columns([0.2, 0.8])
    with col1:
        csv_data = duplicate_names_df.to_csv(index=False)
        st.download_button(
            label="📥 Download (CSV)",
            data=csv_data,
            file_name=f"duplicate_names_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            key="duplicate_names_download"
        )
    st.dataframe(duplicate_names_df, use_container_width=True)
else:
    st.info("No duplicate customer names detected in the dataset.")

# ================= MAIN DASHBOARD SECTIONS =================
st.markdown("---")

# Row 1: Risk Distribution, Collection, and Trend
col1, col2, col3 = st.columns([1.25, 1.6, 1.6])

with col1:
    st.subheader("Risk Distribution")
    risk_dist = filtered_df["Risk_Category"].value_counts()
    fig = px.pie(
        values=risk_dist.values,
        names=risk_dist.index,
        hole=0.45,
        color=risk_dist.index,
        color_discrete_map=RISK_COLORS
    )
    fig.update_traces(
        textposition="inside",
        textinfo="percent+label",
        marker=dict(line=dict(color="#ffffff", width=4)),
        pull=[0.02] * len(risk_dist),
        hovertemplate="<b>%{label}</b><br>Customers: %{value:,}<br>Share: %{percent}<extra></extra>"
    )
    style_chart(fig, height=520, legend_orientation="v")
    fig.update_layout(margin=dict(l=8, r=8, t=24, b=8))
    st.plotly_chart(fig, use_container_width=True)
    risk_reference = pd.DataFrame([
    {
        "Term": "On Track",
        "Criteria": "0-29 days system off",
        "Meaning": "Current or very low risk"
    },
    {
        "Term": "Watchlist",
        "Criteria": "30-89 days system off",
        "Meaning": "Early warning, needs monitoring"
    },
    {
        "Term": "Medium Risk",
        "Criteria": "90-179 days system off",
        "Meaning": "Requires active follow-up"
    },
    {
        "Term": "High Risk",
        "Criteria": "180-364 days system off",
        "Meaning": "Serious concern"
    },
    {
        "Term": "Critical",
        "Criteria": "365+ days system off",
        "Meaning": "Likely default, urgent action"
    },
    {
        "Term": "Completed",
        "Criteria": "State is paid_off",
        "Meaning": "Loan completed"
    },
])

st.markdown("### Risk Category Reference")

st.dataframe(
    risk_reference,
    use_container_width=True,
    hide_index=True,
    height=320,
    column_config={
        "Term": st.column_config.TextColumn(
            "Term",
            width="medium"
        ),
        "Criteria": st.column_config.TextColumn(
            "Criteria",
            width="large"
        ),
        "Meaning": st.column_config.TextColumn(
            "Meaning",
            width="large"
        ),
    }
)

with col2:
    st.subheader("Collection Trend")
    if "Charged until" in filtered_df.columns:
        collection_trend = (
            filtered_df.groupby(filtered_df["Charged until"].dt.date)
            .agg({
                "Monthly_Payment": "sum",
                "Expected_Arrears": "sum",
                "Customer": "count"
            })
            .reset_index()
            .sort_values(by="Charged until")
            .rename(columns={"Customer": "Customer Count"})
        )
        if len(collection_trend) > 0:
            fig = px.scatter(
                collection_trend,
                x="Charged until",
                y="Expected_Arrears",
                size="Customer Count",
                color="Monthly_Payment",
                custom_data=["Customer Count", "Monthly_Payment"],
                color_continuous_scale=["#bbf7d0", "#22c55e", "#14532d"],
                labels={
                    "Charged until": "Due Date",
                    "Expected_Arrears": "Expected Arrears",
                    "Monthly_Payment": "Expected Monthly",
                    "Customer Count": "Customers"
                }
            )
            fig.update_layout(
                xaxis_title="Due Date",
                yaxis_title="Expected Arrears (MWK)",
                xaxis_tickangle=-45,
                coloraxis_colorbar=dict(title="Monthly")
            )
            fig.update_traces(
                marker=dict(line=dict(color="#ffffff", width=1.5), sizemin=6),
                opacity=0.88,
                hovertemplate=(
                    "<b>%{x}</b><br>"
                    "Expected arrears: MK %{y:,.0f}<br>"
                    "Expected monthly: MK %{customdata[1]:,.0f}<br>"
                    "Customers: %{customdata[0]:,}<extra></extra>"
                )
            )
            style_chart(fig, height=480)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No collection trend available for the current selection.")
    else:
        st.info("Charged until date is required for collection trend.")

with col3:
    st.subheader("Due Date Trend")
    if "Charged until" in filtered_df.columns:
        due_source = filtered_df.dropna(subset=["Charged until"]).copy()
        due_source["Due_Date"] = due_source["Charged until"].dt.normalize()
        due_source["Week_Start"] = due_source["Due_Date"] - pd.to_timedelta(due_source["Due_Date"].dt.weekday, unit="D")
        due_source["Weekday"] = due_source["Due_Date"].dt.day_name().str[:3]
        due_counts = (
            due_source.groupby(["Week_Start", "Weekday"])
            .size()
            .reset_index(name="Customer Count")
            .sort_values(by="Week_Start")
        )
        if len(due_counts) > 0:
            fig = px.density_heatmap(
                due_counts,
                x="Week_Start",
                y="Weekday",
                z="Customer Count",
                histfunc="sum",
                color_continuous_scale=["#f0fdf4", "#86efac", "#16a34a", "#14532d"],
                category_orders={"Weekday": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]},
                labels={
                    "Week_Start": "Week Starting",
                    "Weekday": "Due Day",
                    "Customer Count": "Customers"
                }
            )
            fig.update_layout(
                xaxis_title="Week Starting",
                yaxis_title="Due Day",
                xaxis_tickangle=-45,
                coloraxis_colorbar=dict(title="Customers")
            )
            fig.update_traces(
                hovertemplate="<b>%{y}</b><br>Week: %{x}<br>Customers: %{z:,}<extra></extra>"
            )
            style_chart(fig, height=480)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No due-date trend available for the current selection.")
    else:
        st.info("Charged until date is required for due-date trend.")

# Row 2: Additional visual analysis
active_insights = filtered_df[filtered_df["State"].isin(["good", "active"])].copy() if "State" in filtered_df.columns else filtered_df.copy()
top_arrears = pd.DataFrame()
scatter_df = pd.DataFrame()

insight_col3, insight_col4 = st.columns(2)

with insight_col3:
    st.markdown("**Top Contractors by Arrears**")
    if len(active_insights) > 0 and {"Assigned to contractor", "Expected_Arrears"}.issubset(active_insights.columns):
        top_arrears = (
            active_insights
            .groupby("Assigned to contractor", as_index=False)
            .agg(
                Expected_Arrears=("Expected_Arrears", "sum"),
                Customers=("Customer", "count")
            )
            .sort_values("Expected_Arrears", ascending=False)
            .head(10)
        )
        if len(top_arrears) > 0:
            fig = px.bar(
                top_arrears,
                y="Assigned to contractor",
                x="Expected_Arrears",
                orientation="h",
                text="Customers",
                color="Expected_Arrears",
                color_continuous_scale=["#bbf7d0", "#16a34a", "#14532d"],
                labels={
                    "Assigned to contractor": "Contractor",
                    "Expected_Arrears": "Expected Arrears",
                },
            )
            fig.update_traces(
                marker_line_color="#ffffff",
                marker_line_width=1,
                texttemplate="%{text} customers",
                textposition="outside",
                hovertemplate="<b>%{y}</b><br>Arrears: MK %{x:,.0f}<br>%{text} customers<extra></extra>"
            )
            fig.update_layout(yaxis={"categoryorder": "total ascending"}, coloraxis_showscale=False)
            style_chart(fig, height=420)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No contractor arrears available.")
    else:
        st.info("Contractor and arrears fields are required for this view.")

with insight_col4:
    st.markdown("**Customer Risk Scatter**")
    if len(active_insights) > 0 and {"Days system off", "Expected_Arrears", "Risk_Category"}.issubset(active_insights.columns):
        scatter_df = active_insights[
            (active_insights["Expected_Arrears"] > 0) |
            (active_insights["Days system off"] > 0)
        ].copy()
        if len(scatter_df) > 0:
            fig = px.scatter(
                scatter_df,
                x="Days system off",
                y="Expected_Arrears",
                color="Risk_Category",
                size="Weekly_Payment" if "Weekly_Payment" in scatter_df.columns else None,
                hover_name="Customer" if "Customer" in scatter_df.columns else None,
                hover_data=["Assigned to contractor"] if "Assigned to contractor" in scatter_df.columns else None,
                color_discrete_map=RISK_COLORS,
                labels={
                    "Days system off": "Days System Off",
                    "Expected_Arrears": "Expected Arrears",
                    "Risk_Category": "Risk"
                },
            )
            fig.update_traces(
                marker=dict(line=dict(color="#ffffff", width=1), opacity=0.78),
                hovertemplate="<b>%{hovertext}</b><br>Days off: %{x}<br>Arrears: MK %{y:,.0f}<extra></extra>"
            )
            style_chart(fig, height=420)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No overdue customers available for the current selection.")
    else:
        st.info("Days off, arrears, and risk fields are required for this view.")

if len(scatter_df) > 0 and {"Assigned to contractor", "Risk_Category", "Customer", "Expected_Arrears", "Days system off"}.issubset(scatter_df.columns):
    st.markdown("**Contractor Arrears & Risk Table**")
    combined_agg = {
        "Customer": "count",
        "Expected_Arrears": "sum",
        "Days system off": "mean",
    }
    if "Weekly_Payment" in scatter_df.columns:
        combined_agg["Weekly_Payment"] = "sum"

    combined_table = (
        scatter_df
        .groupby(["Assigned to contractor", "Risk_Category"], as_index=False)
        .agg(combined_agg)
        .rename(columns={
            "Assigned to contractor": "Contractor",
            "Risk_Category": "Risk",
            "Customer": "Customers",
            "Expected_Arrears": "Total Arrears",
            "Days system off": "Avg Days System Off",
            "Weekly_Payment": "Weekly Payment"
        })
        .sort_values(["Total Arrears", "Customers"], ascending=[False, False])
    )
    format_rules = {
        "Total Arrears": "MK {:,.0f}",
        "Avg Days System Off": "{:.1f}",
    }
    if "Weekly Payment" in combined_table.columns:
        format_rules["Weekly Payment"] = "MK {:,.0f}"

    st.dataframe(
        combined_table.style.format(format_rules),
        use_container_width=True,
        hide_index=True
    )

# Filtered customer list
st.markdown("---")
st.subheader("📋 Filtered Customers")

if len(filtered_df) > 0:
    display_cols = [
        "Customer",
        "Customer phone numbers",
        "Assigned to contractor",
        "State",
        "Risk_Category",
        "Plan_Type",
        "Payoff amount",
        "Balance",
        "Left to pay",
        "Expected_Arrears",
        "Days system off",
        "Charged until",
        "Location"
    ]
    display_cols = [col for col in display_cols if col in filtered_df.columns]
    
    # Download button
    export_df = make_excel_safe(filtered_df[display_cols].sort_values(by=["Risk_Category", "Charged until"], ascending=[True, True]))
    csv_data = export_df.to_csv(index=False)
    st.download_button(
        label="📥 Download Filtered Customers (CSV)",
        data=csv_data,
        file_name=f"filtered_customers_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mime="text/csv"
    )
    
    st.dataframe(
        filtered_df[display_cols].sort_values(by=["Risk_Category", "Charged until"], ascending=[True, True]).style.format({
            "Payoff amount": "MK {:,.0f}",
            "Balance": "MK {:,.0f}",
            "Left to pay": "MK {:,.0f}",
            "Expected_Arrears": "MK {:,.0f}",
            "Days system off": "{:.0f}"
        }),
        use_container_width=True
    )
else:
    st.info("No customers match the current filter selection.")

# Row 2: Customers per Due Date
st.markdown("---")
st.subheader("📅 Customers by Due Date")

if "Charged until" in filtered_df.columns:
    due_counts = (
        filtered_df.groupby(filtered_df["Charged until"].dt.date)
        .size()
        .reset_index(name="Customer Count")
        .sort_values(by="Charged until")
        .rename(columns={"Charged until": "Due Date"})
    )
    if len(due_counts) > 0:
        due_display = due_counts.copy()
        due_display["Due Date"] = pd.to_datetime(due_display["Due Date"]).dt.strftime("%d %b %Y")
        due_selection = st.dataframe(
            due_display,
            use_container_width=True,
            hide_index=True,
            key="customers_by_due_date_table",
            on_select="rerun",
            selection_mode="single-row"
        )
        selected_due_rows = due_selection.selection.rows
        if selected_due_rows:
            selected_due_date = due_counts.iloc[selected_due_rows[0]]["Due Date"]
            selected_due_customers = filtered_df[
                filtered_df["Charged until"].dt.date == selected_due_date
            ].copy()
            st.caption(f"Customers due on {pd.to_datetime(selected_due_date).strftime('%d %b %Y')}")
            graph_col1, graph_col2 = st.columns(2)

            with graph_col1:
                if "Risk_Category" in selected_due_customers.columns:
                    selected_risk_counts = (
                        selected_due_customers["Risk_Category"]
                        .fillna("Unknown")
                        .value_counts()
                        .rename_axis("Risk")
                        .reset_index(name="Customers")
                    )
                    fig = px.bar(
                        selected_risk_counts,
                        x="Risk",
                        y="Customers",
                        text="Customers",
                        color="Risk",
                        color_discrete_map=RISK_COLORS,
                        labels={"Risk": "Risk", "Customers": "Customers"}
                    )
                    fig.update_traces(
                        marker_line_color="#ffffff",
                        marker_line_width=1,
                        texttemplate="%{y:,}",
                        textposition="outside",
                        hovertemplate="<b>%{x}</b><br>%{y:,} customers<extra></extra>"
                    )
                    fig.update_layout(showlegend=False, bargap=0.28)
                    style_chart(fig, height=320)
                    st.plotly_chart(fig, use_container_width=True)

            with graph_col2:
                if {"Plan_Type", "Monthly_Payment"}.issubset(selected_due_customers.columns):
                    selected_plan_payments = (
                        selected_due_customers
                        .groupby("Plan_Type", as_index=False)
                        .agg(Monthly_Payment=("Monthly_Payment", "sum"))
                        .sort_values("Monthly_Payment", ascending=False)
                    )
                    fig = px.bar(
                        selected_plan_payments,
                        x="Plan_Type",
                        y="Monthly_Payment",
                        text="Monthly_Payment",
                        color_discrete_sequence=[CHART_COLORS["primary"]],
                        labels={"Plan_Type": "Plan", "Monthly_Payment": "Expected Monthly"}
                    )
                    fig.update_traces(
                        marker_line_color="#ffffff",
                        marker_line_width=1,
                        texttemplate="MK %{y:,.0f}",
                        textposition="outside",
                        hovertemplate="<b>%{x}</b><br>MK %{y:,.0f}<extra></extra>"
                    )
                    fig.update_layout(bargap=0.28)
                    style_chart(fig, height=320)
                    st.plotly_chart(fig, use_container_width=True)

            due_customer_cols = [
                "Customer",
                "Assigned to contractor",
                "State",
                "Risk_Category",
                "Plan_Type",
                "Monthly_Payment",
                "Expected_Arrears",
                "Days system off",
                "Location"
            ]
            due_customer_cols = [col for col in due_customer_cols if col in selected_due_customers.columns]
            st.dataframe(
                selected_due_customers[due_customer_cols].sort_values(by="Customer"),
                use_container_width=True,
                hide_index=True
            )
        else:
            st.caption("Select a due-date row to see the customer names for that day.")
    else:
        st.info("No customers available for the selected due date range.")
else:
    st.info("Charged until date is required for due-date grouping.")

# Row 4: Contractor Performance
st.markdown("---")
st.subheader("🏢 Contractor Performance")

contractors_perf = get_all_contractors_collection(filtered_df)
if len(contractors_perf) > 0:
    st.dataframe(
        contractors_perf.style.format({
            "expected_total": "MK {:,.0f}",
            "collected_total": "MK {:,.0f}",
            "financed_total": "MK {:,.0f}",
            "efficiency_percent": "{:.1f}%"
        }),
        use_container_width=True
    )
else:
    st.info("No contractor data available")

# Row 3: Urgent Follow-ups
st.markdown("---")
st.subheader("🔴 Urgent Follow-ups (Top 20)")

urgent = get_urgent_followups(filtered_df, top_n=20)
if len(urgent) > 0:
    st.dataframe(
        urgent.style.format({
            "Weekly_Payment": "MK {:,.0f}",
            "Expected_Arrears": "MK {:,.0f}"
        }),
        use_container_width=True
    )
else:
    st.success("✅ No urgent follow-ups needed!")

# Row 4: Forecast
st.markdown("---")
st.subheader("📊 30-Day Forecast & Scenarios")

col1, col2, col3, col4 = st.columns(4)

with col1:
    st.metric("Daily Projection", f"MK {forecast['daily_avg']:,.0f}")
with col2:
    st.metric("30-Day Forecast", f"MK {forecast['forecast_30_days']:,.0f}")
with col3:
    st.metric("Confidence", f"{forecast['confidence_percent']:.0f}%")
with col4:
    st.metric("Monthly Projection", f"MK {forecast['monthly_projection']:,.0f}")

# Scenario analysis
st.subheader("Scenario Analysis")
scenarios = {}
for scenario in ["conservative", "realistic", "optimistic"]:
    scenarios[scenario] = scenario_analysis(filtered_df, scenario)

scenario_data = []
for scenario, data in scenarios.items():
    scenario_data.append({
        "Scenario": scenario.title(),
        "Weekly": data.get("weekly_collection", 0),
        "Monthly": data.get("monthly_projection", 0)
    })

scenario_df = pd.DataFrame(scenario_data)
st.dataframe(
    scenario_df.style.format({
        "Weekly": "MK {:,.0f}",
        "Monthly": "MK {:,.0f}"
    }),
    use_container_width=True
)

forecast_plot_df = scenario_df.melt(id_vars=["Scenario"], value_vars=["Weekly", "Monthly"],
                                    var_name="Period", value_name="Amount")
fig = px.bar(
    forecast_plot_df,
    x="Scenario",
    y="Amount",
    color="Period",
    color_discrete_sequence=[CHART_COLORS["primary"], CHART_COLORS["accent"]],
    labels={"Amount": "MK", "Scenario": "Scenario", "Period": "Projection"},
)
fig.update_traces(
    marker_line_color="#ffffff",
    marker_line_width=1,
    opacity=0.92,
    hovertemplate="<b>%{x}</b><br>MK %{y:,.0f}<extra></extra>"
)
fig.update_layout(barmode="group", bargap=0.22)
style_chart(fig, height=380)
st.plotly_chart(fig, use_container_width=True)

# Row 5: Critical Cases
st.markdown("---")
st.subheader("⚠️ Critical Cases (180+ days)")

critical = get_critical_cases(filtered_df, 180)
if len(critical) > 0:
    st.warning(f"⚠️ {len(critical)} customers in critical state requiring immediate action")
    st.dataframe(
        critical.style.format({
            "Expected_Arrears": "MK {:,.0f}",
            "Left to pay": "MK {:,.0f}"
        }),
        use_container_width=True
    )
else:
    st.success("✅ No critical cases!")

# ================= EXPORT & DOWNLOAD =================
st.markdown("---")
st.subheader("📥 Export Options")

# Column filter for the filtered data export
if not filtered_df.empty:
    all_cols = filtered_df.columns.tolist()
    preferred_cols = [
        "Customer", "Customer phone numbers", "Assigned to contractor", 
        "State", "Risk_Category", "Plan_Type", "Expected_Arrears", 
        "Days system off", "Charged until", "Left to pay", "Location"
    ]
    default_cols = [c for c in preferred_cols if c in all_cols]
    selected_cols = st.multiselect("Select columns for CSV export", options=all_cols, default=default_cols if default_cols else all_cols)
else:
    selected_cols = []

col1, col2, col3, col4 = st.columns(4)

with col1:
    csv = filtered_df[selected_cols].to_csv(index=False).encode("utf-8") if selected_cols else filtered_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "📊 Download Filtered Data (CSV)",
        csv,
        "recapo_filtered.csv",
        "text/csv"
    )

with col2:
    urgent_csv = get_urgent_followups(filtered_df, 50).to_csv(index=False).encode("utf-8")
    st.download_button(
        "🔴 Download Urgent List (CSV)",
        urgent_csv,
        "recapo_urgent.csv",
        "text/csv"
    )

with col3:
    critical_csv = get_critical_cases(filtered_df, 180).to_csv(index=False).encode("utf-8")
    st.download_button(
        "⚠️ Download Critical Cases (CSV)",
        critical_csv,
        "recapo_critical.csv",
        "text/csv"
    )

with col4:
    excel_dashboard = build_excel_dashboard(
        filtered_df,
        kpi_summary,
        scenario_df,
        contractors_perf,
        get_urgent_followups(filtered_df, 50),
        get_critical_cases(filtered_df, 180),
    )
    st.download_button(
        "Download Excel Dashboard",
        excel_dashboard,
        "recapo_dashboard.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ================= FOOTER =================
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray; font-size: 12px;'>
RECAPO Intelligence System | Recovery & Collection Forecast Engine
</div>
""", unsafe_allow_html=True)
